
# -# *- coding: utf-8 -*-
#Created by Alex Wernle 12/21/20
""" Script is meant to automate creation of Vs30 model plot in Excel
User should copy this script into a folder with Excel xxxxx_SCHOOL_profile, and
final model .rst. User should also change variable fn pathway"""

import pandas as pd
import numpy as np
from pandas import ExcelWriter
from pandas import ExcelFile
from openpyxl import load_workbook
from glob import glob



# Open .rst (does not need to be .txt), read file and split string to list
filename = glob('*.rst')[0]
f = open(filename)
message = f.read()
model_info= message.split() 

#Edits: loop through each .rst file in folder!

# Define new arrays for data
new_array =  []
new_array2 = []


# Fix significant data into arrays before passing to Excel#####################

# Loop through data until key_num(save val and pos),convert to float and append 
# to array
for object in model_info[1:]: #ignore first object
       if len(object) == 2:
           #print(object)
           key_num = object 
           key_loc = model_info.index(key_num)
           break
       object = float(object)
       new_array.append(object)

# Start next loop after key_num location
key_loc = key_loc+1
model_info2 = model_info[key_loc:]

# Find second key_num and key_loc
for object in model_info2:
    if object == key_num:
        #print(object)
        key_loc2 = model_info2.index(object)
        #print(key_loc2)
        break

# Save all dat after second key_num and key_loc
for object in model_info2[key_loc2+1:]:
    object = float(object)
    new_array2.append(object)
  
        
# Reshape arrays into proper matrix dimensions
data = np.array([new_array2])
shape = (int((len(new_array2)/2)),2)
data2= data.reshape(shape)

data3 = np.array([new_array])
shape = (int((len(new_array)/4)),4)
data4= data3.reshape(shape)


##############################################################################
# Edited helper function for Excel
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
     to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """


    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet, set Index= falso to not include numbered rows
    df.to_excel(writer, sheet_name, startrow=startrow,index=False,header=None,
                **to_excel_kwargs)

    # save the workbook
    writer.save()
    
##############################################################################
# Open Excel file, sheet name and write data to respective rows/columns
fn = r'C:\Users\awer490\Desktop\Python_Vs_Model_Test\xxxxx_SCHOOL_profile.xlsx'

df2 = pd.DataFrame(data2)
df3 = pd.DataFrame(data4)


append_df_to_excel(fn,df2,sheet_name='1D Mod & Disp_TN',
                   startcol=2, startrow=(1),
                   truncate_sheet=False)

append_df_to_excel(fn,df3,sheet_name='1D Mod & Disp_TN', 
                   startcol=5, startrow=(1),
                   truncate_sheet=False)

f.close()

#Edits: save excel doc with name, then set dataframes back to 0 

print("Script run successful")

##############################################################################
#script improvement:
    #automatically name excel file
    #empty excel columns
    #