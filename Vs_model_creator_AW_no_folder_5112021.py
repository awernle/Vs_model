# -*- coding: utf-8 -*-

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from glob import glob
import shutil
import subprocess
import os



##############################################################################
# Edited helper function for Excel
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
     to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """


    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet, set Index= falso to not include numbered rows
    df.to_excel(writer, sheet_name, startrow=startrow,index=False,header=None,
                **to_excel_kwargs)

    # save the workbook
    writer.save()
    
##############################################################################
# before starting, make sure these workbook columns are empty except for title:
    # depth	(C)
    # Vs, ms (D)	
    # Vel_Meas (F)
    # Vel_Model	(G)
    # Freq (H)	
    # Coh (I)
# and these columns have the equation ready down an absured number of rows
    # column before "depth" (B)
    # Depth	(J)
    # Coh/max (K)
    
# EACH USER NEEDS TO CHANGE THESE VARIABLES:

pth = 'C:/Users/awer490/Desktop/Python_Vs_Model_Test/Script_edits'
user = 'aw'


# loop through each .rst file in current folder
for filename in glob("*.rst"):
# Open .rst (does not need to be .txt), read file and split string to list
    f = open(filename)
    message = f.read()
    model_info= message.split() 
        #grab the lineid
    lid = filename.split('.')[0]

    # Define new arrays for data
    new_array =  []
    new_array2 = []
    
    
    # Fix significant data into arrays before passing to Excel################
    
    # Loop through data until key_num(save val and pos),convert to float,append 
    # to array
    for object in model_info[1:]: #ignore first object
            if len(object) == 2:
                #print(object)
                key_num = object 
                key_loc = model_info.index(key_num)
                break
            object = float(object)
            new_array.append(object)
    
    # Start next loop after key_num location
    key_loc = key_loc+1
    model_info2 = model_info[key_loc:]
    
    # Find second key_num and key_loc
    for object in model_info2:
        if object == key_num:
            #print(object)
            key_loc2 = model_info2.index(object)
            #print(key_loc2)
            break
    
    # Save all dat after second key_num and key_loc
    for object in model_info2[key_loc2+1:]:
        object = float(object)
        new_array2.append(object)
      
    
    # Reshape arrays into proper matrix dimensions
    data = np.array([new_array2])
    shape = (int((len(new_array2)/2)),2)
    data2= data.reshape(shape)
    
    data3 = np.array([new_array])
    shape = (int((len(new_array)/4)),4)
    data4= data3.reshape(shape)
    
    # before doing anything else, check "vel_model" column is not all 0s
    """if data4[:,1].sum() == 0:
        print('Alas! vel_model column of file ' + lid + ' is all 0s.')
        subprocess.Popen('C:/Program Files (x86)/SeisImager/WaveEq.exe')
        continue"""
  
    
    #move files into folder
    pth = pth + '/' 
    workbook ='xxxxx_SCHOOL_profile.xlsx' 
    src = pth + workbook


    # create a copy of the blank workbook, renamed, to write to
    fn = pth + '/' + lid + '_profile_' + user + '.xlsx' #replace pth with rep?
    shutil.copy(src, fn)
    # also create a powerpoint file with correct name in that folder
    ppoint = 'lineschool_profile.pptx'
    srp = pth + ppoint
    fnp = pth +  '/' + lid + '_profile_' + user + '.pptx'
    shutil.copy(srp, fnp)

            
# # TROUBLESHOOT - this is printing out 14 times, for each file/folder in 
# # Jade_try - want it to just say it once
    
    """   # create a copy of the blank workbook, renamed, to write to
        fn = pth + '/' + lid + '_profile_' + user + '.xlsx'
        shutil.copy(src, fn)
        # also create a named, blank powerpoint pres to copy the profile into
        ppoint = 'lineschool_profile.pptx'
        srp = pth + ppoint
        fnp = pth +  '/' + lid + '_profile_' + user + '.pptx'
        shutil.copy(srp, fnp)"""

    # Open Excel file, sheet name and write data to respective rows/columns
    df2 = pd.DataFrame(data2)
    df3 = pd.DataFrame(data4)
    
    
    append_df_to_excel(fn,df2,sheet_name='1D Mod & Disp_TN',
                       startcol=2, startrow=(1),
                       truncate_sheet=False)
    
    append_df_to_excel(fn,df3,sheet_name='1D Mod & Disp_TN', 
                       startcol=5, startrow=(1),
                       truncate_sheet=False)
    
    #will likely need to create new figure using openpyxl
    
    print("Script run successful for file " + lid)
    f.close()
  
 # when naming path, be sure to change \ to / And have / at end
"""pth = 'C:/Users/jcey490/Desktop/OnePagers/Jade_try/Profile/'
    workbook ='xxxxx_SCHOOL_profile.xlsx'
    src = pth + workbook
    
    # create a copy of the blank workbook, renamed, in correct folder
    for folder in os.listdir('C:/Users/jcey490/Desktop/OnePagers/Jade_try/'):
        fid = folder.split(".")[0]
        if fid == lid:
            scl = folder.split(".")[1]
            rep = 'C:/Users/jcey490/Desktop/OnePagers/Jade_try/' + folder
            fn = rep + '/' + lid + '.' + scl + '.profile_jc.xlsx'
            shutil.copy(src, fn)
            # also create a powerpoint file with correct name in that folder
            spp = pth + 'lineschool_profile_jc.pptx'
            dpp = rep + '/' + lid + '.' + scl + '.profile_jc.pptx'
            shutil.copy(spp,dpp)  """
            # break 
# # TROUBLESHOOT - this is printing out 14 times, for each file/folder in 
# # Jade_try - want it to just say it once 

##############################################################################
